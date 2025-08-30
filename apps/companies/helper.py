import os

from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import status

from apps.companies.models.operation_model import Car, CarOperation
from apps.shared.base_exception_class import CustomValidationError
from apps.shared.send_sms import send_sms


def get_car_operations_data(*args, **kwargs):
    queryset = (
        CarOperation.objects.filter(
            car__branch__company=kwargs.get("company_id"),
            status=CarOperation.OperationStatus.COMPLETED,
        )
        .select_related("car", "driver", "station_branch", "worker", "service")
        .order_by("-id")
    )
    if kwargs.get("car"):
        queryset = queryset.filter(car_id=kwargs.get("car"))
    date_from = kwargs.get("date_from")
    if date_from:
        try:
            date_from = timezone.localtime().strptime(date_from, "%Y-%m-%d").date()
            queryset = queryset.filter(start_time__date__gte=date_from)
        except ValueError:
            raise CustomValidationError(
                message="Invalid date from format",
                status_code=status.HTTP_400_BAD_REQUEST,
            )
    date_to = kwargs.get("date_to")
    if date_to:
        try:
            date_to = timezone.localtime().strptime(date_to, "%Y-%m-%d").date()
            queryset = queryset.filter(start_time__date__lte=date_to)
        except ValueError:
            raise CustomValidationError(
                message="Invalid date to format",
                status_code=status.HTTP_400_BAD_REQUEST,
            )
    return queryset


def send_cash_request_otp(instance):
    message = "كود استلام طلبك النقدي بمقدار {} هو {}".format(
        instance.amount, instance.otp
    )
    send_sms(message, instance.driver.phone_number)


def export_car_operations(*args, **kwargs):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Export"
    center_alignment = Alignment(horizontal="center", vertical="center")
    cars = Car.objects.filter(branch__company=kwargs.get("company_id"))
    if kwargs.get("car"):
        cars = cars.filter(id=kwargs.get("car"))
    is_first_car = True

    for car in cars:
        if not is_first_car:
            empty_row = ["", "", "", "", "", "", ""]
            ws.append(empty_row)
            ws.append(empty_row)
        is_first_car = False

        # Row with car number + "العربية"
        car_number_row = ["", "", "", "", "", str(car), "العربية"]
        ws.append(car_number_row)

        # Style car row: black text + gray background
        bold_font = Font(bold=True, color="000000")  # black text
        car_fill = PatternFill(
            start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
        )  # gray
        row_index = ws.max_row
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_index, column=col)
            cell.font = bold_font
            cell.fill = car_fill
            cell.alignment = center_alignment

        # Header row
        operation_export_header = [
            "التكلفة",
            "rate of fuel cons",
            "الوقود المستهلك",
            "العداد اليومي",
            "آخر عداد",
            "أول عداد",
            "التاريخ",
        ]
        ws.append(operation_export_header)

        # Style header row (blue background + white bold text)
        header_fill = PatternFill(
            start_color="527993", end_color="527993", fill_type="solid"
        )  # blue
        header_font = Font(bold=True, color="FFFFFF")  # white text
        row_index = ws.max_row
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_index, column=col)
            cell.font = header_font
            cell.fill = header_fill

        # Data rows (alternate colors)
        kwargs["car"] = car.id
        car_operations = get_car_operations_data(*args, **kwargs)

        light_fill1 = PatternFill(
            start_color="E5E4E0", end_color="E5E4E0", fill_type="solid"
        )  # white
        light_fill2 = PatternFill(
            start_color="FFFDE7", end_color="FFFDE7", fill_type="solid"
        )  # light yellow

        for idx, row in enumerate(car_operations):
            company_cost = row.company_cost
            cleaned_row = [
                f"{company_cost} جنيه",
                "-",
                row.amount,
                row.fuel_consumption_rate,
                row.car_meter,
                row.car_first_meter,
                row.created.date(),
            ]
            ws.append(cleaned_row)

            # Apply alternating colors
            row_index = ws.max_row
            fill = light_fill1 if idx % 2 == 0 else light_fill2
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_index, column=col)
                cell.fill = fill
                cell.alignment = center_alignment
                col_letter = get_column_letter(col)
                current_width = ws.column_dimensions[col_letter].width or 0

                # Check length of current cell value
                if cell.value:
                    length = len(str(cell.value)) + 2  # Add small padding
                    if length > current_width:
                        ws.column_dimensions[col_letter].width = length

    # Generate filename with timestamp
    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")
    filename = f"export_{timestamp}.xlsx"
    excel_dir = os.path.join(settings.MEDIA_ROOT, "excel_exports")
    os.makedirs(excel_dir, exist_ok=True)
    filepath = os.path.join(excel_dir, filename)

    # Save workbook
    wb.save(filepath)
    return filename
